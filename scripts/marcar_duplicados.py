# -*- coding: utf-8 -*-
"""
Detecta nombres de cliente escritos de varias formas (probables duplicados)
y los marca en AMARILLO dentro de las propias bases .xlsx.

La idea: el Excel manda. El script no "arregla" el nombre, solo lo senala
para que lo corrijas en la fuente; al corregirlo, el dashboard se actualiza.

Uso:
    python scripts/marcar_duplicados.py            # solo reporta (no escribe)
    python scripts/marcar_duplicados.py --marcar   # respalda y marca en amarillo
    python scripts/marcar_duplicados.py --restaurar # deshace desde el respaldo

El marcado se hace por cirugia directa sobre el XML del .xlsx: solo se tocan
styles.xml y el atributo de estilo de las celdas afectadas. Comentarios de
celda, imagenes, dibujos y formatos se conservan intactos.
"""
import json, os, sys, shutil, zipfile, re, argparse, datetime, unicodedata
import xml.etree.ElementTree as ET

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(AQUI)
NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
AMARILLO = "FFFF00"


def norm(s):
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return " ".join(s.upper().replace("\n", " ").split())


def col_letra(i):
    s = ""
    i += 1
    while i:
        i, r = divmod(i - 1, 26)
        s = chr(65 + r) + s
    return s


# ------------------------------------------------------------- deteccion
def leer_clientes(ruta, hoja):
    """Devuelve (indice_hoja_en_libro, [(fila_1based, col_0based, nombre)])."""
    import openpyxl
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    real = next((s for s in wb.sheetnames if norm(s) == norm(hoja)), None)
    if real is None:
        wb.close()
        return None, []
    idx = wb.sheetnames.index(real)
    ws = wb[real]
    filas = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    ic = next((i for i, f in enumerate(filas) if any(norm(c) == "LOTE" for c in f)), None)
    if ic is None:
        return idx, []
    cab = filas[ic]
    jc = next((j for j, c in enumerate(cab) if norm(c) in ("CLIENTE", "NOMBRE DEL CLIENTE")), None)
    if jc is None:
        return idx, []

    out = []
    for i, f in enumerate(filas[ic + 1:], start=ic + 2):   # 1-based para Excel
        v = f[jc] if jc < len(f) else None
        if v is None or not str(v).strip():
            continue
        out.append((i, jc, str(v).strip()))
    return idx, out


def distancia(a, b, tope):
    """Distancia de edicion (Levenshtein), cortando en cuanto supera el tope."""
    if abs(len(a) - len(b)) > tope:
        return tope + 1
    previa = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        actual = [i]
        for j, cb in enumerate(b, 1):
            actual.append(min(previa[j] + 1, actual[j - 1] + 1,
                              previa[j - 1] + (ca != cb)))
        if min(actual) > tope:
            return tope + 1
        previa = actual
    return previa[-1]


def agrupar_parecidos(nombres, tope=2):
    """Agrupa nombres que difieren en 1-2 letras: el patron tipico de errata.

    Se usa distancia de edicion y no similitud porcentual porque en estas bases
    abundan los familiares (mismos apellidos, distinto nombre de pila), que una
    similitud alta confunde con duplicados.
    """
    unicos = sorted({norm(n) for _, _, n in nombres if len(norm(n)) >= 8})
    grupos, usados = [], set()
    for i, a in enumerate(unicos):
        if a in usados:
            continue
        grupo = [(a, 0)]
        for b in unicos[i + 1:]:
            if b in usados:
                continue
            d = distancia(a, b, tope)
            if d <= tope:
                grupo.append((b, d))
                usados.add(b)
        if len(grupo) > 1:
            usados.add(a)
            grupos.append(grupo)
    return grupos


# ------------------------------------------------------------- marcado xml
def estilo_amarillo(xml_estilos):
    """Anade (si falta) un relleno amarillo y un cellXf que lo usa. Devuelve (xml, indice)."""
    ET.register_namespace("", NS)
    root = ET.fromstring(xml_estilos)

    fills = root.find(f"{{{NS}}}fills")
    idx_fill = None
    for k, f in enumerate(fills.findall(f"{{{NS}}}fill")):
        pf = f.find(f"{{{NS}}}patternFill")
        if pf is not None and pf.get("patternType") == "solid":
            fg = pf.find(f"{{{NS}}}fgColor")
            if fg is not None and (fg.get("rgb") or "").upper().endswith(AMARILLO):
                idx_fill = k
                break
    if idx_fill is None:
        f = ET.SubElement(fills, f"{{{NS}}}fill")
        pf = ET.SubElement(f, f"{{{NS}}}patternFill", {"patternType": "solid"})
        ET.SubElement(pf, f"{{{NS}}}fgColor", {"rgb": "FF" + AMARILLO})
        ET.SubElement(pf, f"{{{NS}}}bgColor", {"indexed": "64"})
        idx_fill = len(fills.findall(f"{{{NS}}}fill")) - 1
        fills.set("count", str(idx_fill + 1))

    xfs = root.find(f"{{{NS}}}cellXfs")
    idx_xf = None
    for k, xf in enumerate(xfs.findall(f"{{{NS}}}xf")):
        if xf.get("fillId") == str(idx_fill) and xf.get("applyFill") == "1":
            idx_xf = k
            break
    if idx_xf is None:
        xf = ET.SubElement(xfs, f"{{{NS}}}xf", {
            "numFmtId": "0", "fontId": "0", "fillId": str(idx_fill),
            "borderId": "0", "xfId": "0", "applyFill": "1"})
        idx_xf = len(xfs.findall(f"{{{NS}}}xf")) - 1
        xfs.set("count", str(idx_xf + 1))

    return ET.tostring(root, encoding="UTF-8", xml_declaration=True).decode("utf-8"), idx_xf


def ruta_de_hoja(zin, nombre_hoja):
    """Traduce el nombre visible de la hoja a su xml dentro del zip.

    No se puede asumir que la 3a hoja sea sheet3.xml: el orden visible y la
    numeracion interna son independientes. Hay que seguir la relacion r:id.
    """
    wb = ET.fromstring(zin.read("xl/workbook.xml"))
    rels = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
    RNS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    destino = {r.get("Id"): r.get("Target") for r in rels}

    for s in wb.find(f"{{{NS}}}sheets"):
        if norm(s.get("name")) == norm(nombre_hoja):
            t = destino.get(s.get(f"{{{RNS}}}id"), "")
            if not t:
                return None
            t = t.lstrip("/")
            return t if t.startswith("xl/") else "xl/" + t
    return None


def aplicar(ruta, nombre_hoja, celdas):
    """Pinta de amarillo las celdas indicadas, conservando todo lo demas.

    Solo se reescriben styles.xml y el xml de la hoja; el resto de las entradas
    del zip (comentarios, imagenes, dibujos) se copia tal cual.
    """
    zin = zipfile.ZipFile(ruta, "r")
    hoja_path = ruta_de_hoja(zin, nombre_hoja)
    if hoja_path is None or hoja_path not in zin.namelist():
        zin.close()
        return 0, "no ubico la hoja '%s' dentro del archivo" % nombre_hoja

    estilos, idx_xf = estilo_amarillo(zin.read("xl/styles.xml").decode("utf-8"))
    hoja = zin.read(hoja_path).decode("utf-8")
    objetivo = set(celdas)
    tocadas = [0]

    def sub_celda(m):
        ref = m.group("ref")
        if ref not in objetivo:
            return m.group(0)
        tocadas[0] += 1
        etiqueta = m.group(0)
        if re.search(r'\ss="\d+"', etiqueta):
            return re.sub(r'\ss="\d+"', ' s="%d"' % idx_xf, etiqueta, count=1)
        return etiqueta[:2] + ' s="%d"' % idx_xf + etiqueta[2:]

    hoja = re.sub(r'<c r="(?P<ref>[A-Z]+\d+)"[^>]*?/?>', sub_celda, hoja)

    tmp = ruta + ".tmp"
    zout = zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED)
    for item in zin.infolist():
        datos = zin.read(item.filename)
        if item.filename == "xl/styles.xml":
            datos = estilos.encode("utf-8")
        elif item.filename == hoja_path:
            datos = hoja.encode("utf-8")
        zout.writestr(item, datos)
    zout.close()
    zin.close()
    os.replace(tmp, ruta)
    return tocadas[0], None


def restaurar(carpeta):
    """Deshace el marcado devolviendo las bases desde el respaldo mas reciente."""
    resp = sorted(d for d in os.listdir(carpeta)
                  if d.startswith("_respaldo_") and os.path.isdir(os.path.join(carpeta, d)))
    if not resp:
        print("No hay respaldos que restaurar.")
        return
    origen = os.path.join(carpeta, resp[-1])
    n = 0
    for f in os.listdir(origen):
        shutil.copy2(os.path.join(origen, f), os.path.join(carpeta, f))
        n += 1
    print("Restaurados %d archivos desde %s" % (n, resp[-1]))


# ------------------------------------------------------------- principal
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--marcar", action="store_true", help="escribe el amarillo en las bases")
    ap.add_argument("--restaurar", action="store_true",
                    help="deshace el marcado desde el respaldo mas reciente")
    ap.add_argument("--tope", type=int, default=2,
                    help="maxima diferencia de letras para considerar errata (1 = muy estricto)")
    args = ap.parse_args()

    reglas = json.load(open(os.path.join(AQUI, "reglas.json"), encoding="utf-8"))
    carpeta = os.path.normpath(os.path.join(RAIZ, reglas["carpeta_bases"]))

    if args.restaurar:
        restaurar(carpeta)
        return

    if args.marcar:
        sello = datetime.date.today().isoformat()
        resp = os.path.join(carpeta, "_respaldo_%s" % sello)
        os.makedirs(resp, exist_ok=True)

    total_grupos = total_celdas = 0
    for p in reglas["proyectos"]:
        ruta = next((os.path.join(carpeta, f) for f in os.listdir(carpeta)
                     if norm(f) == norm(p["archivo"])), None)
        if not ruta:
            continue

        for hoja in p["hojas"]:
            idx, clientes = leer_clientes(ruta, hoja)
            if idx is None or not clientes:
                continue
            grupos = agrupar_parecidos(clientes, args.tope)
            if not grupos:
                continue

            sospechosos = {n for g in grupos for n, _ in g}
            celdas = [(f, c, v) for f, c, v in clientes if norm(v) in sospechosos]
            refs = {col_letra(c) + str(f) for f, c, v in celdas}

            print("\n%s  ·  hoja %s" % (p["nombre"], hoja))
            for g in grupos:
                total_grupos += 1
                base = g[0][0]
                otros = "   /   ".join("%s (%d letra%s)" % (n, d, "" if d == 1 else "s")
                                       for n, d in g[1:])
                print("   ~ %s   /   %s" % (base, otros))
            total_celdas += len(refs)

            if args.marcar:
                destino = os.path.join(resp, os.path.basename(ruta))
                if not os.path.exists(destino):
                    shutil.copy2(ruta, destino)
                n, err = aplicar(ruta, hoja, refs)
                print("     -> %s" % (err if err else "marcadas %d celdas" % n))

    print("\n" + "=" * 62)
    print("  %d grupos de nombres parecidos, %d celdas" % (total_grupos, total_celdas))
    if not args.marcar:
        print("  (solo reporte: no se escribio nada. Usa --marcar para pintarlas)")
    else:
        print("  respaldo en: %s" % resp)
        print("  para deshacer:  python scripts/marcar_duplicados.py --restaurar")
    print("=" * 62)


if __name__ == "__main__":
    main()
