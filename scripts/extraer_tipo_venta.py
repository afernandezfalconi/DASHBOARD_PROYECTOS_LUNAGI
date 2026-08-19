# -*- coding: utf-8 -*-
"""
Saca el tipo de venta (contado o credito) de cada lote y lo deja en
tipo_venta.json, para que el Sheet lo traiga sin perder ediciones manuales.

Las bases no lo guardan igual:

  Familia ACTIVOS Y FINIQUITOS -> columna TIPO DE VENTA / TIPO DE COMPRA,
      con valores limpios CREDITO o CONTADO.

  Familia MENSUALIDADES (ZOI, GUAYACAN, NEXUS, INTEROCEANICO) -> columna
      STATUS, que MEZCLA el tipo de venta con el estado de pago: ademas de
      CREDITO y CONTADO trae LIQUIDADO, que solo dice "ya termino de pagar"
      y no si fue de contado o a credito.

      Para esos se deduce: si la venta traia mas de una mensualidad fue a
      credito; si no, fue de contado. Con las 252 filas LIQUIDADO de las
      bases actuales la deduccion resuelve el 100% (153 credito, 99 contado).

    python scripts/extraer_tipo_venta.py
"""
import json, os, sys

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instalalo con:  pip install openpyxl")

import auditar
from auditar import norm, num, texto

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(AQUI)

COL_TIPO = ("TIPO DE VENTA", "TIPO DE COMPRA", "STATUS", "TIPO")
CREDITO, CONTADO, COMISION = "Crédito", "Contado", "Comisión"


def clasificar(valor, mensualidades, enganche, precio):
    v = norm(valor)
    if not v or v in ("NA", "N/A", "-"):
        return ""
    if v.startswith("CONTADO"):
        return CONTADO
    if v.startswith("CREDIT") or v.startswith("CRÉDIT"):
        return CREDITO
    if v.startswith("COMISION"):
        return COMISION
    if v.startswith("LIQUID"):                 # estado de pago, no tipo de venta
        if mensualidades > 1:
            return CREDITO
        if precio and abs(enganche - precio) < 1:
            return CONTADO
        return CONTADO
    return ""


def main():
    reglas = json.load(open(os.path.join(AQUI, "reglas.json"), encoding="utf-8"))
    carpeta = os.path.normpath(os.path.join(RAIZ, reglas["carpeta_bases"]))
    alias = reglas["columnas"]

    salida, resumen = {}, []
    for p in reglas["proyectos"]:
        ruta = auditar.buscar_archivo(carpeta, p["archivo"])
        if not ruta:
            continue
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)

        datos, cuenta = {}, {CREDITO: 0, CONTADO: 0, COMISION: 0}
        columna = None
        for nombre in p["hojas"]:
            real = next((s for s in wb.sheetnames if norm(s) == norm(nombre)), None)
            if real is None:
                continue
            filas = [list(f) for f in wb[real].iter_rows(values_only=True)]
            ic = auditar.hallar_cabecera(filas, alias["lote"])
            if ic is None:
                continue
            cab = filas[ic]
            mapa = auditar.mapear_columnas(cab, alias)
            jt = next((j for j, c in enumerate(cab) if norm(c) in COL_TIPO), None)
            jm = next((j for j, c in enumerate(cab) if norm(c) == "MENSUALIDADES"), None)
            if jt is None or "lote" not in mapa or "mza" not in mapa:
                continue
            columna = norm(cab[jt])

            for f in filas[ic + 1:]:
                def v(j):
                    return f[j] if j is not None and j < len(f) else None

                lote = texto(v(mapa["lote"]))
                if not lote:
                    continue
                t = clasificar(v(jt), num(v(jm)),
                               num(v(mapa.get("enganche"))), num(v(mapa.get("monto"))))
                if t:
                    datos[(texto(v(mapa["mza"])), lote)] = t
                    cuenta[t] += 1
        wb.close()

        salida[p["nombre"]] = {"%s|%s" % k: v for k, v in datos.items()}
        resumen.append((p["nombre"], columna or "SIN COLUMNA",
                        cuenta[CREDITO], cuenta[CONTADO], cuenta[COMISION]))

    destino = os.path.join(RAIZ, "tipo_venta.json")
    with open(destino, "w", encoding="utf-8") as fh:
        json.dump(salida, fh, ensure_ascii=False, separators=(",", ":"))

    print("%-22s %-16s %9s %9s %9s" % ("PROYECTO", "COLUMNA", "CRÉDITO", "CONTADO", "COMISIÓN"))
    print("-" * 72)
    for n, c, cr, co, cm in resumen:
        print("%-22s %-16s %9d %9d %9d" % (n, c[:16], cr, co, cm))
    print("-" * 72)
    print("%-22s %-16s %9d %9d %9d" % ("TOTAL", "",
                                       sum(r[2] for r in resumen),
                                       sum(r[3] for r in resumen),
                                       sum(r[4] for r in resumen)))
    print("\n-> %s" % destino)


if __name__ == "__main__":
    main()
