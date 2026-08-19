# -*- coding: utf-8 -*-
"""
Saca la superficie (m2) de cada lote desde las bases originales y la deja en
m2.json, para que el Sheet la pueda jalar sin perder las ediciones manuales.

Busca en dos lugares:
  1. La columna M2 de la hoja principal (familia ACTIVOS Y FINIQUITOS)
  2. La columna AREA de la pestana de PRECIOS, emparejando por MZA + LOTE
     (asi se recuperan ZOI y GUAYACAN, que no traen M2 en su hoja principal)

    python scripts/extraer_m2.py
"""
import json, os, sys, unicodedata

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instalalo con:  pip install openpyxl")

import auditar
from auditar import norm, num, texto

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(AQUI)

COLS_M2 = ("M2", "M²", "AREA", "AREA M2", "SUPERFICIE", "MTS2", "METROS")
# pestanas de precios que traen AREA por mza/lote
HOJAS_PRECIO = ("PRECIOS", "TABLA DE PRECIOS", "LISTA DE PRECIOS")


def cabecera_y_cols(filas):
    """Localiza la fila de encabezado y las columnas mza / lote / m2."""
    for i, f in enumerate(filas[:30]):
        if not any(norm(c) == "LOTE" for c in f):
            continue
        jl = next(j for j, c in enumerate(f) if norm(c) == "LOTE")
        jm = next((j for j, c in enumerate(f) if norm(c) in ("MZA", "MZA.", "MANZANA", "MZ")), None)
        j2 = next((j for j, c in enumerate(f) if norm(c) in COLS_M2), None)
        if jm is not None:
            return i, jm, jl, j2
    return None, None, None, None


def leer(ws):
    """Devuelve {(mza, lote): m2} de una hoja, si trae los datos."""
    filas = [list(f) for f in ws.iter_rows(values_only=True)]
    ic, jm, jl, j2 = cabecera_y_cols(filas)
    if ic is None or j2 is None:
        return {}
    out = {}
    for f in filas[ic + 1:]:
        if jl >= len(f) or jm >= len(f):
            continue
        lote = texto(f[jl])
        if not lote:
            continue
        v = num(f[j2]) if j2 < len(f) else 0
        if v > 0:
            out[(texto(f[jm]), lote)] = round(v, 2)
    return out


def main():
    reglas = json.load(open(os.path.join(AQUI, "reglas.json"), encoding="utf-8"))
    carpeta = os.path.normpath(os.path.join(RAIZ, reglas["carpeta_bases"]))

    salida, resumen = {}, []
    for p in reglas["proyectos"]:
        ruta = auditar.buscar_archivo(carpeta, p["archivo"])
        if not ruta:
            continue
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)

        datos, fuente = {}, []
        for nombre in p["hojas"]:                      # 1) hoja principal
            real = next((s for s in wb.sheetnames if norm(s) == norm(nombre)), None)
            if real:
                d = leer(wb[real])
                if d:
                    datos.update(d)
                    fuente.append("hoja principal")

        if not datos:                                  # 2) pestana de precios
            for s in wb.sheetnames:
                if any(norm(s).startswith(h) for h in HOJAS_PRECIO):
                    d = leer(wb[s])
                    if d:
                        datos.update(d)
                        fuente.append("pestaña '%s'" % s.strip())
                        break
        wb.close()

        salida[p["nombre"]] = {"%s|%s" % k: v for k, v in datos.items()}
        resumen.append((p["nombre"], len(datos), sum(datos.values()),
                        ", ".join(dict.fromkeys(fuente)) or "SIN DATO"))

    destino = os.path.join(RAIZ, "m2.json")
    with open(destino, "w", encoding="utf-8") as fh:
        json.dump(salida, fh, ensure_ascii=False, separators=(",", ":"))

    print("%-22s %8s %14s   %s" % ("PROYECTO", "LOTES", "SUMA M2", "DE DONDE"))
    print("-" * 78)
    for n, c, s, f in resumen:
        print("%-22s %8d %14s   %s" % (n, c, format(s, ",.2f"), f))
    print("-" * 78)
    print("%-22s %8d %14s" % ("TOTAL", sum(r[1] for r in resumen),
                              format(sum(r[2] for r in resumen), ",.2f")))
    print("\n-> %s" % destino)


if __name__ == "__main__":
    main()
