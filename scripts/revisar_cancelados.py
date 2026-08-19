# -*- coding: utf-8 -*-
"""
Lista los lotes en los que el Sheet quedo con los datos de una fila CANCELADA
en lugar de los de la fila activa.

Por que existen: la auditoria original suponia que la fila que aparece despues
del marcador CANCELADOS era el estado mas reciente y debia sustituir a la de
arriba. Es al reves: lo que va debajo del marcador esta cancelado. El Sheet se
sembro con la regla equivocada, asi que esos lotes muestran al comprador que se
dio de baja y su anticipo, en vez del comprador real y lo que lleva pagado.

Genera dos archivos junto al repositorio:
    revisar_cancelados.csv    para abrir en Excel y corregir a mano
    correcciones.json         mismo contenido, por si se automatiza despues

    python scripts/revisar_cancelados.py
"""
import csv, json, os, sys

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instalalo con:  pip install openpyxl")

import auditar
from auditar import norm, num, texto

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(AQUI)

_VACIOS, _PREFIJOS = set(), []


def limpiar_cliente(nombre):
    """'DISPONIBLE', '.', '0' y demas no son compradores: son lote libre."""
    n = norm(nombre)
    if n in _VACIOS or any(n.startswith(p) for p in _PREFIJOS):
        return ""
    return nombre


def aprender_homologacion(reglas, afectados):
    """Deduce como se renombraron los clientes, comparando el Sheet con las bases.

    El usuario homologo nombres a mano en el Sheet (HECTOR -> LUNAMAR, los PG a
    PG INMOBILIARIA...). Las bases siguen crudas. Si al corregir un lote
    cancelado copiaramos el nombre de la base, desharíamos ese trabajo.

    CUIDADO con el sesgo: en los lotes AFECTADOS el Sheet trae el nombre del
    comprador cancelado, asi que compararlos daria "NORMA -> MUSS", que no es
    una homologacion sino justo el error a corregir. Por eso se aprende solo de
    los lotes sanos, que se reciben en `afectados` para excluirlos.
    """
    ruta = os.path.join(RAIZ, "datos.json")
    if not os.path.exists(ruta):
        return {}
    datos = json.load(open(ruta, encoding="utf-8"))
    sheet = {}
    for p in datos.get("projects", []):
        for l in p["lotes"]:
            sheet[(p["nombre"], l["mza"], l["lote"])] = l["cliente"]

    carpeta = os.path.normpath(os.path.join(RAIZ, reglas["carpeta_bases"]))
    alias = reglas["columnas"]
    cortes = reglas.get("marcadores_de_corte", {}).get("textos", ["CANCELADOS"])

    votos = {}
    for p in reglas["proyectos"]:
        ruta_x = auditar.buscar_archivo(carpeta, p["archivo"])
        if not ruta_x:
            continue
        wb = openpyxl.load_workbook(ruta_x, read_only=True, data_only=True)
        for nombre in p["hojas"]:
            real = next((s for s in wb.sheetnames if norm(s) == norm(nombre)), None)
            if real is None:
                continue
            activas, _ = filas_por_bloque(wb[real], alias, cortes)
            for k, dat in activas.items():
                if (p["nombre"], k[0], k[1]) in afectados:
                    continue                      # ahi el Sheet trae el cancelado
                crudo = dat["cliente"]
                puesto = sheet.get((p["nombre"], k[0], k[1]))
                if not crudo or not puesto or norm(crudo) == norm(puesto):
                    continue
                votos.setdefault(norm(crudo), {}).setdefault(puesto, 0)
                votos[norm(crudo)][puesto] += 1
        wb.close()

    # Solo se adoptan las equivalencias que se repiten. Una sola aparicion suele
    # ser una reasignacion puntual del lote (cambio de comprador), no un cambio
    # de nombre: aplicarla a otros lotes pondria al cliente equivocado.
    firmes, dudosas = {}, {}
    for k, v in votos.items():
        nombre, veces = max(v.items(), key=lambda x: x[1])
        (firmes if veces >= 2 else dudosas)[k] = nombre
    return firmes, dudosas


def filas_por_bloque(ws, alias, cortes):
    """Separa la hoja en (activas, canceladas), ambas {(mza, lote): datos}."""
    filas = [list(f) for f in ws.iter_rows(values_only=True)]
    ic = auditar.hallar_cabecera(filas, alias["lote"])
    if ic is None:
        return {}, {}
    cab = filas[ic]
    mapa = auditar.mapear_columnas(cab, alias)
    if "lote" not in mapa or "mza" not in mapa:
        return {}, {}

    corte_pag = mapa.get("corte_pagos")
    pagos = (auditar.columnas_de_pago(cab, corte_pag, alias.get("excluir_de_pagos", []))
             if corte_pag is not None else [])

    def leer(f):
        def v(rol):
            j = mapa.get(rol)
            return f[j] if j is not None and j < len(f) else None
        abonado = num(v("abonado"))
        enganche = num(v("enganche"))
        mensual = sum(num(f[j]) for j in pagos if j < len(f))
        return {
            "cliente": limpiar_cliente(texto(v("cliente"))),
            "monto": round(num(v("monto")), 2),
            "enganche": round(enganche, 2),
            "ingreso": round(abonado if abonado else enganche + mensual, 2),
        }

    activas, canceladas, corte = {}, {}, False
    for f in filas[ic + 1:]:
        if auditar.es_marcador_de_corte(f, cortes):
            corte = True
            continue
        lote = texto(f[mapa["lote"]] if mapa["lote"] < len(f) else None)
        if not lote:
            continue
        k = (texto(f[mapa["mza"]] if mapa["mza"] < len(f) else None), lote)
        (canceladas if corte else activas)[k] = leer(f)
    return activas, canceladas


def main():
    reglas = json.load(open(os.path.join(AQUI, "reglas.json"), encoding="utf-8"))
    carpeta = os.path.normpath(os.path.join(RAIZ, reglas["carpeta_bases"]))
    alias = reglas["columnas"]
    cortes = reglas.get("marcadores_de_corte", {}).get("textos", ["CANCELADOS"])

    global _VACIOS, _PREFIJOS
    _VACIOS = {norm(x) for x in reglas.get("cliente_vacio", [])}
    _PREFIJOS = [norm(x) for x in reglas.get("cliente_vacio_prefijo", [])]

    casos = []
    for p in reglas["proyectos"]:
        ruta = auditar.buscar_archivo(carpeta, p["archivo"])
        if not ruta:
            continue
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        for nombre in p["hojas"]:
            real = next((s for s in wb.sheetnames if norm(s) == norm(nombre)), None)
            if real is None:
                continue
            activas, canceladas = filas_por_bloque(wb[real], alias, cortes)
            for k, cancel in canceladas.items():
                if k not in activas:
                    continue                       # solo existe cancelado
                act = activas[k]
                casos.append({
                    "proyecto": p["nombre"], "manzana": k[0], "lote": k[1],
                    "cliente_correcto": act["cliente"],
                    "cliente_en_la_base": act["cliente"],
                    "cliente_cancelado": cancel["cliente"],
                    "monto_correcto": act["monto"],
                    "enganche_correcto": act["enganche"],
                    "ingreso_correcto": act["ingreso"],
                    "ingreso_cancelado": cancel["ingreso"],
                    "diferencia": round(act["ingreso"] - cancel["ingreso"], 2),
                })
        wb.close()

    # La homologacion se aprende DESPUES de saber cuales son los afectados, para
    # no aprender el propio error como si fuera un cambio de nombre intencional.
    afectados = {(c["proyecto"], c["manzana"], c["lote"]) for c in casos}
    homol, dudosas = aprender_homologacion(reglas, afectados)
    for k, v in reglas.get("homologacion_clientes", {}).items():
        if not k.startswith("_"):
            homol[norm(k)] = v          # forzadas a mano en reglas.json
    renombrados = 0
    for c in casos:
        nuevo = homol.get(norm(c["cliente_en_la_base"]))
        if nuevo and norm(nuevo) != norm(c["cliente_correcto"]):
            c["cliente_correcto"] = nuevo
            renombrados += 1
    print("Homologaciones firmes (se repiten): %d" % len(homol))
    print("Correcciones que adoptan tu nombre homologado: %d" % renombrados)
    usadas = {norm(c["cliente_en_la_base"]) for c in casos}
    pasadas = [(k, v) for k, v in dudosas.items() if k in usadas]
    if pasadas:
        print()
        print("Equivalencias vistas UNA sola vez, NO aplicadas (revisalas):")
        for k, v in pasadas:
            print("   %-38s -> %s" % (k[:38], v[:38]))

    casos.sort(key=lambda c: -abs(c["diferencia"]))

    csv_out = os.path.join(RAIZ, "revisar_cancelados.csv")
    with open(csv_out, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(casos[0].keys()) if casos else ["proyecto"])
        w.writeheader()
        w.writerows(casos)

    json_out = os.path.join(RAIZ, "correcciones.json")
    with open(json_out, "w", encoding="utf-8") as fh:
        json.dump(casos, fh, ensure_ascii=False, separators=(",", ":"))

    dif = sum(c["diferencia"] for c in casos)
    print("=" * 74)
    print("  %d lotes con datos de una fila CANCELADA en vez de la activa" % len(casos))
    print("  Efecto neto en ingresos si se corrigen: %s%s" %
          ("+" if dif >= 0 else "-", format(abs(dif), ",.2f")))
    print("=" * 74)
    print()
    print("%-20s %5s %5s  %-28s %14s" % ("PROYECTO", "MZA", "LOTE", "COMPRADOR REAL", "DIFERENCIA"))
    print("-" * 78)
    for c in casos[:20]:
        print("%-20s %5s %5s  %-28s %14s" %
              (c["proyecto"][:20], c["manzana"], c["lote"],
               c["cliente_correcto"][:28], format(c["diferencia"], ",.2f")))
    if len(casos) > 20:
        print("  ...y %d mas (estan todos en el csv)" % (len(casos) - 20))
    print()
    print("-> %s" % csv_out)
    print("-> %s" % json_out)


if __name__ == "__main__":
    main()
