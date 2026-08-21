# -*- coding: utf-8 -*-
"""
Crea AUDITORIA_LOTES.xlsx: UN solo archivo con los 2,202 lotes de los 18
proyectos, para trabajar todo desde ahi en vez de abrir 18 bases distintas.

    python scripts/generar_maestro.py            # lo crea (no pisa si ya existe)
    python scripts/generar_maestro.py --recrear  # lo rehace desde las bases

OJO: --recrear descarta las ediciones manuales del maestro y vuelve a leer las
bases originales. Antes de hacerlo guarda una copia con fecha.
"""
import json, os, sys, shutil, argparse, datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Falta openpyxl. Instalalo con:  pip install openpyxl")

import auditar
from marcar_duplicados import agrupar_parecidos

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(AQUI)

# orden visible de la hoja LOTES (M2 y TIPO VENTA se llenan luego desde las bases)
CAB = ["PROYECTO", "MANZANA", "LOTE", "M2", "CLIENTE", "TIPO VENTA", "ESTATUS",
       "MONTO", "ENGANCHE", "INGRESO", "CATEGORIA", "NOTA"]
ANCHOS = [24, 11, 10, 9, 44, 13, 19, 15, 15, 15, 27, 34]

AZUL = "1F4E79"
AMARILLO = "FFFF00"
COLOR_ESTATUS = {
    "DISPONIBLE": "DDEBF7",
    "APARTADO": "FCE4D6",
    "VENDIDO": "E2EFDA",
    "VENDIDO SIN DATO": "FFF2CC",
}
CATEGORIAS = ["Inmobiliaria socia", "Administracion Cristhian", "Uso interno / comunal",
              "Anotacion sin cliente", "Cliente sin dato capturado"]


def hoja_leeme(wb):
    ws = wb.create_sheet("LEEME", 0)
    ws.column_dimensions["A"].width = 104
    lineas = [
        ("AUDITORIA DE LOTES - LUNA GI", True),
        ("", False),
        ("Este archivo es la FUENTE DE VERDAD del dashboard vivo.", True),
        ("https://afernandezfalconi.github.io/DASHBOARD_PROYECTOS_LUNAGI/", False),
        ("", False),
        ("COMO USARLO", True),
        ("1. Edita lo que necesites en la hoja LOTES.", False),
        ("2. Guarda y cierra este archivo.", False),
        ("3. Doble clic en actualizar.cmd (esta junto a este archivo).", False),
        ("4. En un minuto el dashboard ya muestra tus cambios.", False),
        ("", False),
        ("QUE PUEDES EDITAR", True),
        ("CLIENTE    el nombre del comprador. Dejalo VACIO si el lote esta libre.", False),
        ("ESTATUS    elige de la lista: DISPONIBLE / APARTADO / VENDIDO / VENDIDO SIN DATO", False),
        ("MONTO      precio del lote", False),
        ("ENGANCHE   enganche registrado", False),
        ("INGRESO    lo efectivamente cobrado. ES LA COLUMNA QUE SUMA EL DASHBOARD.", False),
        ("CATEGORIA  solo aplica a los VENDIDO SIN DATO", False),
        ("NOTA       texto libre; aparece en el dashboard cuando el lote no tiene cliente", False),
        ("", False),
        ("REGLAS QUE CONVIENE RESPETAR", True),
        ("- PROYECTO, MANZANA y LOTE identifican la fila. No los dupliques.", False),
        ("- Si pones ESTATUS = DISPONIBLE, deja CLIENTE e INGRESO vacios o en cero.", False),
        ("- Para agregar un lote nuevo, escribe una fila mas al final. Se toma igual.", False),
        ("- Para quitar un lote, borra la fila completa.", False),
        ("", False),
        ("CELDAS EN AMARILLO", True),
        ("Nombres que difieren en 1 o 2 letras de otro del mismo proyecto: posibles", False),
        ("erratas de captura. Revisalos y corrigelos; el amarillo es solo una senal.", False),
        ("", False),
        ("LAS 18 BASES ORIGINALES", True),
        ("Siguen en la carpeta Bd_s PXM y SC como materia prima, pero el dashboard", False),
        ("YA NO las lee. Cuando llegue la base de un mes nuevo:", False),
        ("     python scripts\\importar_bases.py        (muestra que cambiaria)", False),
        ("     python scripts\\importar_bases.py --aplicar", False),
    ]
    for i, (txt, fuerte) in enumerate(lineas, 1):
        c = ws.cell(row=i, column=1, value=txt)
        c.font = Font(bold=fuerte, size=12 if fuerte and i == 1 else 11,
                      color=AZUL if fuerte else "000000")
    ws.sheet_view.showGridLines = False
    return ws


def hoja_resumen(wb, proyectos, ultima_fila):
    ws = wb.create_sheet("RESUMEN")
    enc = ["PROYECTO", "TOTAL", "DISPONIBLES", "APARTADOS", "VENDIDOS",
           "SIN INGRESO", "INGRESOS"]
    for j, t in enumerate(enc, 1):
        c = ws.cell(row=1, column=j, value=t)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 26
    for j in range(2, 8):
        ws.column_dimensions[get_column_letter(j)].width = 15

    R = "LOTES!$A$2:$A$%d" % ultima_fila      # PROYECTO
    E = "LOTES!$G$2:$G$%d" % ultima_fila      # ESTATUS
    I = "LOTES!$J$2:$J$%d" % ultima_fila      # INGRESO
    for i, p in enumerate(proyectos, 2):
        ws.cell(row=i, column=1, value=p["nombre"])
        a = "$A%d" % i
        ws.cell(row=i, column=2, value="=COUNTIF(%s,%s)" % (R, a))
        ws.cell(row=i, column=3, value='=COUNTIFS(%s,%s,%s,"DISPONIBLE")' % (R, a, E))
        ws.cell(row=i, column=4, value='=COUNTIFS(%s,%s,%s,"APARTADO")' % (R, a, E))
        ws.cell(row=i, column=5, value='=COUNTIFS(%s,%s,%s,"VENDIDO")+COUNTIFS(%s,%s,%s,"VENDIDO SIN DATO")'
                % (R, a, E, R, a, E))
        ws.cell(row=i, column=6, value='=COUNTIFS(%s,%s,%s,"VENDIDO SIN DATO")' % (R, a, E))
        c = ws.cell(row=i, column=7, value="=SUMIF(%s,%s,%s)" % (R, a, I))
        c.number_format = '"$"#,##0.00'

    f = len(proyectos) + 2
    ws.cell(row=f, column=1, value="TOTAL").font = Font(bold=True)
    for j in range(2, 8):
        c = ws.cell(row=f, column=j, value="=SUM(%s2:%s%d)" % (get_column_letter(j), get_column_letter(j), f - 1))
        c.font = Font(bold=True)
        if j == 7:
            c.number_format = '"$"#,##0.00'
    ws.freeze_panes = "A2"
    ws.cell(row=f + 2, column=1,
            value="Calculado con formulas sobre la hoja LOTES: cambia solo al editar.").font = \
        Font(italic=True, size=10, color="808080")
    return ws


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--recrear", action="store_true",
                    help="rehace el maestro desde las bases, descartando ediciones")
    args = ap.parse_args()

    reglas = json.load(open(os.path.join(AQUI, "reglas.json"), encoding="utf-8"))
    destino = os.path.normpath(os.path.join(RAIZ, reglas.get("archivo_maestro", "AUDITORIA_LOTES.xlsx")))

    if os.path.exists(destino) and not args.recrear:
        sys.exit("Ya existe %s\nSi de verdad quieres rehacerlo desde las bases (y perder\n"
                 "las ediciones manuales), usa:  python scripts/generar_maestro.py --recrear"
                 % os.path.basename(destino))

    if os.path.exists(destino):
        copia = destino.replace(".xlsx", " (antes de recrear %s).xlsx" % datetime.date.today().isoformat())
        shutil.copy2(destino, copia)
        print("Copia previa: %s" % os.path.basename(copia))

    diag = []
    crudos = auditar.desde_bases(reglas, diag)
    proyectos, _ = auditar.agregar(crudos, reglas)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hoja_leeme(wb)
    ws = wb.create_sheet("LOTES")

    for j, t in enumerate(CAB, 1):
        c = ws.cell(row=1, column=j, value=t)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(j)].width = ANCHOS[j - 1]

    borde = Border(bottom=Side(style="thin", color="D9D9D9"))
    fila = 2
    sospechosos = set()
    for p in proyectos:
        nombres = [(0, 0, l["cliente"]) for l in p["lotes"] if l["cliente"]]
        for g in agrupar_parecidos(nombres, 2):
            for n, _ in g:
                sospechosos.add((p["nombre"], n))

    for p in proyectos:
        for l in p["lotes"]:
            et = auditar.ESTATUS_A_TEXTO[l["estatus"]]
            vals = [p["nombre"], l["mza"], l["lote"], l.get("m2", ""), l["cliente"],
                    l.get("tipo", ""), et, l["monto"], l["enganche"], l["ingreso"],
                    l.get("categoria", ""), l.get("nota", "")]
            for j, v in enumerate(vals, 1):
                c = ws.cell(row=fila, column=j, value=v)
                c.border = borde
                if j in (8, 9, 10):
                    c.number_format = '#,##0.00'
                if j == 7:
                    c.fill = PatternFill("solid", fgColor=COLOR_ESTATUS[et])
                    c.alignment = Alignment(horizontal="center")
                if j == 5 and l["cliente"] and (p["nombre"], auditar.norm(l["cliente"])) in sospechosos:
                    c.fill = PatternFill("solid", fgColor=AMARILLO)
            fila += 1

    ultima = fila - 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(CAB)), ultima)

    dv = DataValidation(type="list", allow_blank=False,
                        formula1='"%s"' % ",".join(auditar.ESTATUS_A_TEXTO.values()),
                        showErrorMessage=True)
    dv.error = "Usa uno de la lista: DISPONIBLE, APARTADO, VENDIDO, VENDIDO SIN DATO"
    dv.errorTitle = "Estatus no valido"
    ws.add_data_validation(dv)
    dv.add("G2:G%d" % (ultima + 400))

    dvc = DataValidation(type="list", allow_blank=True,
                         formula1='"%s"' % ",".join(CATEGORIAS), showErrorMessage=False)
    ws.add_data_validation(dvc)
    dvc.add("K2:K%d" % (ultima + 400))

    hoja_resumen(wb, proyectos, ultima)
    wb.save(destino)

    print("\n" + "=" * 62)
    print("  Maestro creado: %s" % os.path.basename(destino))
    print("  %d lotes de %d proyectos" % (ultima - 1, len(proyectos)))
    print("  %d nombres marcados en amarillo" % len(sospechosos))
    print("  Hojas: LEEME (instrucciones) | LOTES (edita aqui) | RESUMEN (formulas)")
    print("=" * 62)


if __name__ == "__main__":
    main()
