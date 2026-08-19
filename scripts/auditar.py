# -*- coding: utf-8 -*-
"""
Auditoria de lotes - LUNA GI
Lee las bases .xlsx y genera datos.json para el dashboard.

NO edites datos.json a mano: se regenera con este script.
Las reglas de homologacion viven en scripts/reglas.json
"""
import json, os, sys, unicodedata, datetime, argparse

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instalalo con:  pip install openpyxl")

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(AQUI)


# ---------------------------------------------------------------- utilidades
def norm(s):
    """Mayusculas, sin acentos, sin espacios extra."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return " ".join(s.upper().replace("\n", " ").split())


def num(v):
    """Convierte a float lo que se pueda; si no, 0.0"""
    if v is None or isinstance(v, bool):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, (datetime.datetime, datetime.date)):
        return 0.0
    s = str(v).strip().replace("$", "").replace(",", "").replace(" ", "")
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def texto(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return ""
    return " ".join(str(v).split())


def r2(x):
    # redondeo estandar de Python (bancario), igual que la auditoria original
    return round(x, 2)


# ---------------------------------------------------------------- estructura
def hallar_cabecera(filas, alias_lote):
    """Devuelve el indice de la fila que contiene la columna LOTE."""
    objetivo = {norm(a) for a in alias_lote}
    for i, fila in enumerate(filas[:25]):
        for c in fila:
            if norm(c) in objetivo:
                return i
    return None


def mapear_columnas(cabecera, cols):
    """Empareja cada rol de columna con su indice, por texto exacto del encabezado."""
    mapa = {}
    for rol, alias in cols.items():
        objetivo = {norm(a) for a in alias}
        for j, c in enumerate(cabecera):
            if norm(c) in objetivo:
                mapa.setdefault(rol, j)
    return mapa


def columnas_de_pago(cabecera, desde, excluir):
    """Columnas de MONTO mensual.

    Van despues del corte (MONTO A FINIQUITAR / A FINANCIAR), tienen encabezado,
    no son de FECHA y no son columnas de resumen (TOTAL ABONADO, SALDO A PAGAR...).
    Sumar las de resumen inflaba los ingresos: eran saldos, no pagos.
    """
    fuera = {norm(x) for x in excluir}
    out = []
    for j in range(desde + 1, len(cabecera)):
        h = norm(cabecera[j])
        if not h or h.startswith("FECHA") or h in fuera:
            continue
        out.append(j)
    return out


def es_marcador_de_corte(fila, textos):
    """Fila divisoria: CANCELADOS o CAMBIO DE FRACCIONAMIENTO.

    Todo lo que viene despues esta cancelado (o el cliente se llevo su dinero a
    otro fraccionamiento). No es inventario activo ni ingreso historico: se
    ignora por completo. La fila valida de ese lote es la de ARRIBA del corte.
    """
    objetivo = {norm(t) for t in textos}
    return any(norm(c) in objetivo for c in fila)


def clasificar_cliente(cliente, clasif):
    """Agrupa un 'vendido sin ingreso' segun quien aparece como cliente."""
    n = norm(cliente)
    for cat in clasif.get("categorias", []):
        if n in {norm(x) for x in cat.get("exactos", [])}:
            return cat["nombre"]
        if any(norm(x) in n for x in cat.get("contiene", [])):
            return cat["nombre"]
    return clasif.get("categoria_por_defecto", "Cliente sin dato capturado")


# ---------------------------------------------------------------- auditoria
def auditar_hoja(ws, reglas, diag):
    filas = [list(f) for f in ws.iter_rows(values_only=True)]
    if not filas:
        return []

    cols = reglas["columnas"]
    ic = hallar_cabecera(filas, cols["lote"])
    if ic is None:
        diag.append("   !! no se hallo fila de encabezado")
        return []

    cabecera = filas[ic]
    mapa = mapear_columnas(cabecera, cols)
    if "lote" not in mapa or "mza" not in mapa:
        diag.append("   !! falta columna MZA o LOTE")
        return []

    corte = mapa.get("corte_pagos")
    pagos = (columnas_de_pago(cabecera, corte, cols.get("excluir_de_pagos", []))
             if corte is not None else [])

    diag.append(
        "   cols: mza=%s lote=%s cliente=%s monto=%s enganche=%s abonado=%s corte=%s pagos=%d"
        % (mapa.get("mza"), mapa.get("lote"), mapa.get("cliente"), mapa.get("monto"),
           mapa.get("enganche"), mapa.get("abonado"), corte, len(pagos))
    )

    clasif = reglas.get("clasificacion", {})
    vacios = {norm(x) for x in reglas["cliente_vacio"]}
    prefijos = [norm(x) for x in reglas["cliente_vacio_prefijo"]]

    cortes = reglas.get("marcadores_de_corte", {}).get("textos", ["CANCELADOS"])

    lotes = {}          # (mza, lote) -> registro
    for fila in filas[ic + 1:]:
        if not any(c is not None and str(c).strip() for c in fila):
            continue
        if es_marcador_de_corte(fila, cortes):
            break                          # de aqui para abajo esta cancelado

        def val(rol):
            j = mapa.get(rol)
            return fila[j] if j is not None and j < len(fila) else None

        mza, lote = texto(val("mza")), texto(val("lote"))
        if not lote:
            continue

        cliente = texto(val("cliente"))
        nc = norm(cliente)
        if nc in vacios or any(nc.startswith(p) for p in prefijos):
            cliente = ""

        monto = num(val("monto"))
        enganche = num(val("enganche"))
        abonado = num(val("abonado"))

        mensual = 0.0
        for j in pagos:
            if j < len(fila):
                mensual += num(fila[j])

        ingreso = abonado if abonado else (enganche + mensual)

        # --- estatus del contrato (cancelado / apartado) ---
        ec = norm(val("estatus"))
        ec = norm(clasif.get("correcciones_estatus", {}).get(ec, ec))
        nota, categoria = "", None

        if cliente and ec in {norm(x) for x in clasif.get("estatus_cancelado", [])}:
            nota = "Cancelado: " + cliente     # el lote vuelve a inventario
            cliente = ""

        if not cliente:
            estatus = "disponible"
            ingreso = 0.0
        elif ec in {norm(x) for x in clasif.get("estatus_apartado", [])}:
            estatus = "apartado"
        elif monto or enganche or mensual or abonado:
            estatus = "vendido"
        else:
            estatus = "vendido_sin_dato"
            ingreso = 0.0
            categoria = clasificar_cliente(cliente, clasif)

        reg = {
            "mza": mza, "lote": lote, "cliente": cliente, "estatus": estatus,
            "monto": r2(monto) if monto else None,
            "enganche": r2(enganche) if enganche else None,
            "ingreso": r2(ingreso),
        }
        if categoria:
            reg["categoria"] = categoria
        if nota:
            reg["nota"] = nota
        lotes[(mza, lote)] = reg

    return list(lotes.values())


def buscar_archivo(carpeta, nombre):
    """Empareja por nombre exacto, tolerando acentos compuestos/descompuestos."""
    objetivo = norm(nombre)
    for f in sorted(os.listdir(carpeta)):
        if not f.endswith(".xlsx") or f.startswith("~$"):
            continue
        if norm(f) == objetivo:
            return os.path.join(carpeta, f)
    return None


# ------------------------------------------------------- fuentes de datos
ESTATUS_A_TEXTO = {
    "disponible": "DISPONIBLE",
    "apartado": "APARTADO",
    "vendido": "VENDIDO",
    "vendido_sin_dato": "VENDIDO SIN DATO",
}
TEXTO_A_ESTATUS = {norm(v): k for k, v in ESTATUS_A_TEXTO.items()}
COLS_MAESTRO = ["PROYECTO", "MANZANA", "LOTE", "CLIENTE", "ESTATUS",
                "MONTO", "ENGANCHE", "INGRESO", "CATEGORIA", "NOTA"]


def desde_bases(reglas, diag):
    """Lee los 18 .xlsx originales. Es el camino para sembrar el maestro."""
    carpeta = os.path.normpath(os.path.join(RAIZ, reglas["carpeta_bases"]))
    if not os.path.isdir(carpeta):
        sys.exit("No encuentro la carpeta de bases: %s" % carpeta)

    proyectos = []
    for p in reglas["proyectos"]:
        ruta = buscar_archivo(carpeta, p["archivo"])
        diag.append("%s  <-  %s" % (p["nombre"], os.path.basename(ruta) if ruta else "?? NO ENCONTRADO"))
        if not ruta:
            continue

        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        lotes = []
        for nombre_hoja in p["hojas"]:
            real = next((s for s in wb.sheetnames if norm(s) == norm(nombre_hoja)), None)
            if real is None:
                diag.append("   !! sin hoja '%s' (hay: %s)" % (nombre_hoja, wb.sheetnames))
                continue
            diag.append("   hoja '%s'" % real)
            lotes += auditar_hoja(wb[real], reglas, diag)
        wb.close()
        proyectos.append({"nombre": p["nombre"], "lotes": lotes})
    return proyectos


def desde_maestro(ruta, diag):
    """Lee el archivo unico de trabajo (AUDITORIA_LOTES.xlsx)."""
    if not os.path.exists(ruta):
        sys.exit("No existe el maestro: %s\n"
                 "Generalo con:  python scripts/generar_maestro.py" % ruta)

    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    hoja = next((s for s in wb.sheetnames if norm(s) == "LOTES"), None)
    if hoja is None:
        wb.close()
        sys.exit("El maestro no tiene la hoja LOTES")

    filas = [list(f) for f in wb[hoja].iter_rows(values_only=True)]
    wb.close()

    ic = next((i for i, f in enumerate(filas) if norm(f[0] if f else "") == "PROYECTO"), None)
    if ic is None:
        sys.exit("No encuentro el encabezado (fila con PROYECTO) en la hoja LOTES")
    cab = [norm(c) for c in filas[ic]]
    idx = {c: cab.index(c) for c in COLS_MAESTRO if c in cab}
    if "M2" in cab:
        idx["M2"] = cab.index("M2")              # opcional
    if "TIPO VENTA" in cab:
        idx["TIPO VENTA"] = cab.index("TIPO VENTA")   # opcional
    faltan = [c for c in COLS_MAESTRO if c not in idx]
    if faltan:
        sys.exit("Al maestro le faltan columnas: %s" % ", ".join(faltan))

    porp, orden = {}, []
    saltadas = 0
    for f in filas[ic + 1:]:
        def v(c):
            j = idx[c]
            return f[j] if j < len(f) else None

        proyecto = texto(v("PROYECTO"))
        lote = texto(v("LOTE"))
        if not proyecto or not lote:
            if any(x is not None and str(x).strip() for x in f):
                saltadas += 1
            continue

        est = TEXTO_A_ESTATUS.get(norm(v("ESTATUS")))
        if est is None:
            diag.append("   !! estatus no reconocido '%s' en %s mza %s lote %s -> se omite la fila"
                        % (texto(v("ESTATUS")), proyecto, texto(v("MANZANA")), lote))
            saltadas += 1
            continue

        cliente = texto(v("CLIENTE"))
        monto, enganche = num(v("MONTO")), num(v("ENGANCHE"))
        reg = {
            "mza": texto(v("MANZANA")), "lote": lote, "cliente": cliente, "estatus": est,
            "monto": r2(monto) if monto else None,
            "enganche": r2(enganche) if enganche else None,
            "ingreso": r2(num(v("INGRESO"))),
        }
        if est == "vendido_sin_dato":
            reg["categoria"] = texto(v("CATEGORIA")) or "Cliente sin dato capturado"
        if "M2" in idx:
            sup = num(v("M2"))
            if sup:
                reg["m2"] = r2(sup)
        if "TIPO VENTA" in idx and texto(v("TIPO VENTA")):
            reg["tipo"] = texto(v("TIPO VENTA"))
        if texto(v("NOTA")):
            reg["nota"] = texto(v("NOTA"))

        if proyecto not in porp:
            porp[proyecto] = []
            orden.append(proyecto)
        porp[proyecto].append(reg)

    diag.append("maestro: %d proyectos, %d lotes%s"
                % (len(orden), sum(len(v) for v in porp.values()),
                   ", %d filas omitidas" % saltadas if saltadas else ""))
    return [{"nombre": n, "lotes": porp[n]} for n in orden]


def es_venta(l):
    return l["estatus"] in ("vendido", "vendido_sin_dato")


def precio_por_m2(proyectos):
    """Promedio $/m2 usando solo lotes vendidos que traigan monto y superficie."""
    dinero = metros = 0.0
    for p in proyectos:
        for l in p["lotes"]:
            if l.get("m2") and l.get("monto") and l["estatus"] == "vendido":
                dinero += l["monto"]
                metros += l["m2"]
    return r2(dinero / metros) if metros else 0


def agregar(proyectos):
    """Calcula totales y alertas a partir de la lista de proyectos con lotes."""
    alertas = []
    salida = []
    for p in proyectos:
        lotes = p["lotes"]
        for l in lotes:
            if l["estatus"] == "vendido_sin_dato":
                alertas.append({"proyecto": p["nombre"], "mza": l["mza"], "lote": l["lote"],
                                "cliente": l["cliente"], "categoria": l.get("categoria", "")})
        salida.append({
            "nombre": p["nombre"],
            "total": len(lotes),
            "disponibles": sum(1 for l in lotes if l["estatus"] == "disponible"),
            "apartados": sum(1 for l in lotes if l["estatus"] == "apartado"),
            "vendidos": sum(1 for l in lotes if l["estatus"] in ("vendido", "vendido_sin_dato")),
            "vendidos_con_ingreso": sum(1 for l in lotes if l["estatus"] == "vendido"),
            "vendidos_sin_ingreso": sum(1 for l in lotes if l["estatus"] == "vendido_sin_dato"),
            "ingresos": r2(sum(l["ingreso"] for l in lotes)),
            "m2_total": r2(sum(l.get("m2") or 0 for l in lotes)),
            "m2_disponible": r2(sum(l.get("m2") or 0 for l in lotes if l["estatus"] == "disponible")),
            "m2_vendido": r2(sum(l.get("m2") or 0 for l in lotes
                                 if l["estatus"] in ("vendido", "vendido_sin_dato"))),
            "lotes_con_m2": sum(1 for l in lotes if l.get("m2")),
            "contado": sum(1 for l in lotes if es_venta(l) and norm(l.get("tipo")) == "CONTADO"),
            "credito": sum(1 for l in lotes if es_venta(l) and norm(l.get("tipo")) == "CREDITO"),
            "ingresos_contado": r2(sum(l["ingreso"] for l in lotes
                                       if es_venta(l) and norm(l.get("tipo")) == "CONTADO")),
            "ingresos_credito": r2(sum(l["ingreso"] for l in lotes
                                       if es_venta(l) and norm(l.get("tipo")) == "CREDITO")),
            "lotes": lotes,
        })
    salida.sort(key=lambda p: -p["total"])

    def clave(a):
        def n(x):
            return (0, int(x)) if str(x).strip().isdigit() else (1, str(x))
        return (a["proyecto"], n(a["mza"]), n(a["lote"]))

    alertas.sort(key=clave)   # mismo orden lea de donde lea
    return salida, alertas


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--salida", default=os.path.join(RAIZ, "datos.json"))
    ap.add_argument("--silencio", action="store_true")
    ap.add_argument("--solo-datos", dest="solo_datos", action="store_true",
                    help="genera solo el json, sin regenerar index.html")
    ap.add_argument("--desde", choices=["maestro", "bases"], default=None,
                    help="de donde leer; por defecto, lo que diga reglas.json")
    args = ap.parse_args()

    reglas = json.load(open(os.path.join(AQUI, "reglas.json"), encoding="utf-8"))
    fuente = args.desde or reglas.get("fuente", "bases")
    diag = []

    if fuente == "maestro":
        ruta = os.path.normpath(os.path.join(RAIZ, reglas.get("archivo_maestro", "AUDITORIA_LOTES.xlsx")))
        diag.append("FUENTE: maestro -> %s" % ruta)
        crudos = desde_maestro(ruta, diag)
    else:
        diag.append("FUENTE: las 18 bases originales")
        crudos = desde_bases(reglas, diag)

    proyectos, alertas = agregar(crudos)
    datos = {
        "generated": datetime.date.today().isoformat(),
        "grand": {
            "total": sum(p["total"] for p in proyectos),
            "available": sum(p["disponibles"] for p in proyectos),
            "reserved": sum(p["apartados"] for p in proyectos),
            "sold": sum(p["vendidos"] for p in proyectos),
            "sold_with_data": sum(p["vendidos_con_ingreso"] for p in proyectos),
            "sold_no_data": sum(p["vendidos_sin_ingreso"] for p in proyectos),
            "income": r2(sum(p["ingresos"] for p in proyectos)),
            "m2_total": r2(sum(p["m2_total"] for p in proyectos)),
            "m2_available": r2(sum(p["m2_disponible"] for p in proyectos)),
            "m2_sold": r2(sum(p["m2_vendido"] for p in proyectos)),
            "lots_with_m2": sum(p["lotes_con_m2"] for p in proyectos),
            "price_per_m2": precio_por_m2(proyectos),
            "cash_lots": sum(p["contado"] for p in proyectos),
            "credit_lots": sum(p["credito"] for p in proyectos),
            "cash_income": r2(sum(p["ingresos_contado"] for p in proyectos)),
            "credit_income": r2(sum(p["ingresos_credito"] for p in proyectos)),
        },
        "projects": proyectos,
        "alerts": alertas,
    }

    with open(args.salida, "w", encoding="utf-8") as fh:
        json.dump(datos, fh, ensure_ascii=False, separators=(",", ":"))

    # ---- dashboard: plantilla + datos incrustados (queda autocontenido) ----
    if not args.solo_datos:
        plantilla = os.path.join(AQUI, "plantilla.html")
        if os.path.exists(plantilla):
            html = open(plantilla, encoding="utf-8").read()
            crudo = json.dumps(datos, ensure_ascii=False, separators=(",", ":"))
            # evita que un "</script>" dentro de un dato corte el bloque
            crudo = crudo.replace("</", "<\\/")
            ini, fin = "/*__DATOS_AQUI__*/", "/*__FIN__*/"
            a, b = html.index(ini), html.index(fin)
            html = html[:a + len(ini)] + crudo + html[b:]
            destino = os.path.join(RAIZ, "index.html")
            with open(destino, "w", encoding="utf-8") as fh:
                fh.write(html)
            diag.append("dashboard -> %s" % destino)
        else:
            diag.append("!! falta scripts/plantilla.html; no se regenero index.html")

    if not args.silencio:
        print("\n".join(diag))
    g = datos["grand"]
    print("\n" + "=" * 62)
    print("  Lotes %d | Disponibles %d | Vendidos %d" % (g["total"], g["available"], g["sold"]))
    print("  Con ingreso %d | Sin dato %d" % (g["sold_with_data"], g["sold_no_data"]))
    print("  Ingresos $%s" % format(g["income"], ",.2f"))
    print("  -> %s" % args.salida)
    print("=" * 62)


if __name__ == "__main__":
    main()
